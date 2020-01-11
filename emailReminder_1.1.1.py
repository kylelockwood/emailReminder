#! python3
# Sends emails based on criteria
# This script is set to run weekly in windows Task Scheduler

import openpyxl, smtplib, sys
import datetime as dt


def login_email(smtpObj, user, userPass):
    # Log in to email account
    print('Connecting to email server... ', end='', flush=True)
    smtpObj.ehlo()
    smtpObj.starttls()
    smtpObj.login(user, userPass)
    print('Done')


def find_next_date(sheet):
    # Get the range of dates from the sheet
    dateRange = []
    for c in range (56, 83): #Columns BD through CE
        dateVal = sheet.cell(row=2, column=c).value
        if dateVal != None:
            dateRange.append(sheet.cell(row=2, column=c).value)
    # Find the column with the upcoming date
    today = dt.datetime.today()
    #today = dt.datetime(2020, 1, 22, 0, 0) # test date
    futureDates = [date
                    for date in dateRange
                    if date >= today]
    nextDate = min(futureDates, default=None)
    return nextDate


def get_call_time(sheet):
    callTimes = {}
    for r in range(2,11):
        role = sheet.cell(row=r, column=1).value
        time = sheet.cell(row=r, column=2).value
        callTimes[role] = time
    return callTimes


def get_xl_data(sheet, searchDate):
    # Find the column with the searchDate
    for c in range(56, 83): #Columns BD through CE
        dateCol = sheet.cell(row=2, column=c).value
        if dateCol == searchDate:
            dateCol = int(c)
            break
    # Get volunteers information on scheduled date
    volEmails = {}
    volPhones = {}
    volRoles  = {}
    for r in range(3, 30):
        serviceDate = sheet.cell(row=r, column=dateCol).value
        if serviceDate != None:
            name  = sheet.cell(row=r, column=55).value
            email = sheet.cell(row=r, column=54).value
            phone = sheet.cell(row=r, column=84).value
            role  = sheet.cell(row=r, column=dateCol + 1).value
            volEmails[name] = email
            volPhones[name] = phone
            volRoles[name]  = role
    return volEmails, volPhones, volRoles


def send_emails(printDate, volEmails, volRoles, callTimes):
    statusLog = []
    if len(volEmails) == 0:
        message = 'No volunteers scheduled for %s.' % printDate
        statusLog.append(message)
        print(message)
    else:
        for name, email in volEmails.items():
            role = volRoles.get(name)
            callTime = callTimes.get(role)
            body = (f'Subject: Reminder: You are scheduled for \'{role}\' on {printDate}.\n'
                    f'Hi {name},\n\n'
                    f'This is a reminder that you are scheduled for \'{role}\' at The Oregon Community on {printDate} at {callTime}.\n'
                    'For further schedule details, please visit www.theoregoncommunity.com/calendar\n\n'
                    'See you Sunday!\n\n'
                    'Kyle Lockwood\n'
                    'Director of Operations\n'
                    'The Oregon Community')
            if email == None:
                message = '%s does not have an associated email address' % name
                statusLog.append(message)
                print(message)
            else:
                message = 'Sending remind email to %s... ' % email
                print(message,end='', flush=True)
                #sendmailStatus = smtpObj.sendmail(user, email, body) # turn on to send emails
                sendmailStatus = {} # turn off when sending emails
                if sendmailStatus != {}:
                    status = 'Incomplete'
                    message = message + status
                    statusLog.append(message)
                    print(status)
                    message = 'There was a problem sending email to %s: %s' % (email, sendmailStatus)
                    statusLog.append(message)
                    print(message)
                else:
                    status = 'Completed'
                    message = message + status
                    statusLog.append(message)
                    print(status)
    return statusLog


def email_status(user, searchDate, statusLog):
    # Send completed status email to user
    print(f'Sending status email to {user}... ', end='', flush=True)
    now = dt.datetime.now().strftime('%I:%M:%S%p on %x')
    body = 'Subject: TOC Reminder Emails Status for %s\n' % searchDate
    body = body + '\nReminder emails sent for %s at %s:\n\n' % (searchDate, now)
    for line in statusLog:
        body = body + line + '\n'
    body = body + '\nEnd of status'
    sendmailStatus = smtpObj.sendmail(user, user, body)
    if sendmailStatus != {}:
        print('Incomplete')
        print(f'There was a problem sending the status email to {user}: {sendmailStatus}')
    else:
        print('Completed')



path = 'E:\\Google Drive\\TOC\\'
workbook = path + 'TOC Team Schedule Jan-Mar20.xlsm'
ws = 'role_sort'
user = sys.argv[1]
userPass = sys.argv[2]
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)


# Get spreadsheet data
print(f'Collecting excel data... ', end='', flush=True)
wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True) 
sheet = wb[ws]
nextDate = find_next_date(sheet)
volEmails, volPhones, volRoles = get_xl_data(sheet, nextDate)
callTimes = get_call_time(wb['call_time'])

# Reformat date for email
nextDate = nextDate.strftime('%x')
print('Done')

# Login to email server
login_email(smtpObj, user, userPass)

# Send reminder emails
statusLog = send_emails(nextDate, volEmails, volRoles, callTimes) # This is turned off in the function while testing

# TODO Send text messages to those with phonenumbers
    # TODO skip if they have no phone number provided
    # TODO Update statusLog

# Send status email
email_status(user, nextDate, statusLog)
 
smtpObj.quit()
print('Done')